import threading
from flask import Blueprint, jsonify, request, current_app
from app.models.job_run import JobRun
from app.models.conflict_log import ConflictLog
from app.models.headcount_record import HeadcountRecord
from app.models.medic_record import MedicRecord
from app.models.periodic_record import PeriodicRecord
from app.models.record_change_log import RecordChangeLog
from app.extensions import db
from app.routes.auth import login_required, write_required, biometrics_required
from app import limiter
import os
import json
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook

api_bp = Blueprint("api", __name__)

# Prevents simultaneous long-running job execution across all workers
_task_lock = threading.Lock()

TASK_MAP = {
    "inbox":          "app.services.email_processor",
    "headcount":      "app.services.headcount",
    "booking_alerts": "app.services.booking_alerts",
    "pdf":            "app.services.pdf_processor",
    "medic":          "app.services.medic_reconciler",
}

STATUS_FILTERS = {"overdue", "due_soon", "up_to_date", "new_hires"}
PERIODIC_FIELD_TO_HEADER = {
    "employee_number": "EmployeeID",
    "personnel_name": "Personnel Names",
    "gender": "Gender",
    "role": "Role",
    "department": "Department",
    "subarea": "SubArea",
    "ps_group": "PSGroup",
    "grade": "Grade",
    "base": "Base",
    "date_done": "DateDone",
    "next_due": "NextDue",
    "update_status": "UpdateStatus",
    "update_date": "UpdateDate",
    "days_remaining": "DaysRemaining",
    "status_flag": "StatusFlag",
    "fallback_used": "FallbackUsed",
    "source_file": "SourceFile",
    "form_type": "FormType",
    "surname": "Surname",
    "clinic": "Clinic",
    "needs_review": "NeedsReview",
    "review_reason": "ReviewReason",
}
HEADCOUNT_FIELD_TO_HEADER = {
    "personnel_name": "Personnel Names",
    "employee_number": "Pers.No.",
    "department": "Personnel Area",
    "personnel_subarea": "Personnel Subarea",
    "position": "Position",
    "section": "Section",
    "sub_section": "Sub Section",
    "ps_group": "PS group",
    "age": "Age of employee",
    "flag_type": "Flag Type",
    "flagged_on": "Flagged On",
    "source_file": "Source File",
    "resolved": "Resolved",
    "job_run_id": "Job Run ID",
}
INTEGER_FIELDS = {"days_remaining", "age", "job_run_id", "employee_number"}
BOOLEAN_FIELDS = {"resolved"}
DATE_TEXT_FIELDS = {"date_done", "next_due", "update_date"}


def _clean_text(value):
    if pd.isna(value):
        return None
    text = str(value).strip()
    return text or None


def _safe_int(value):
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    try:
        return int(float(str(value).strip()))
    except (ValueError, TypeError):
        return None


def _coalesce_text(row, *col_names):
    """Return the first non-null text value from the given column names."""
    for col in col_names:
        val = _clean_text(row.get(col))
        if val is not None:
            return val
    return None


def _coalesce_int(row, *col_names):
    """Return the first non-null integer value from the given column names."""
    for col in col_names:
        val = _safe_int(row.get(col))
        if val is not None:
            return val
    return None


def _normalise_lookup(value):
    return (_clean_text(value) or "").strip().lower()


def _format_date_value(value):
    if pd.isna(value) or value in ("", None):
        return None
    parsed = pd.to_datetime(value, errors="coerce")
    if pd.isna(parsed):
        return _clean_text(value)
    return parsed.strftime("%d-%b-%Y")


def _periodic_status_category(days_remaining, warning_days):
    if days_remaining is None:
        return "unknown"
    if days_remaining < 0:
        return "overdue"
    if days_remaining <= warning_days:
        return "due_soon"
    return "up_to_date"


def _coerce_value(field, value):
    if field in INTEGER_FIELDS:
        if value in ("", None):
            return None
        return int(value)
    if field in BOOLEAN_FIELDS:
        if isinstance(value, bool):
            return value
        if value in (None, ""):
            return False
        return str(value).strip().lower() in {"true", "1", "yes", "y"}
    if field in DATE_TEXT_FIELDS:
        return _format_date_value(value)
    if field == "flagged_on":
        if value in ("", None):
            return None
        parsed = pd.to_datetime(value, errors="coerce")
        if pd.isna(parsed):
            raise ValueError("Invalid flagged-on date")
        return parsed.to_pydatetime()
    return _clean_text(value)


def _coerce_workbook_value(field, value):
    if field == "flagged_on":
        if value in ("", None):
            return None
        parsed = pd.to_datetime(value, errors="coerce")
        if pd.isna(parsed):
            return _clean_text(value)
        return parsed.strftime("%d-%b-%Y")
    if field in DATE_TEXT_FIELDS:
        return _format_date_value(value)
    if field in BOOLEAN_FIELDS:
        return "Yes" if value else "No"
    return value


def _serialise_periodic_record(record):
    warning_days = current_app.config.get("DAYS_WARNING", 30) if current_app else 30
    data = record.to_dict()
    data["source"] = "periodic"
    data["status_category"] = _periodic_status_category(record.days_remaining, warning_days)
    return data


def _serialise_headcount_record(record):
    return {
        "source": "headcount",
        "id": record.id,
        "employee_number": record.employee_number,
        "personnel_name": record.personnel_name,
        "department": record.department,
        "personnel_subarea": record.personnel_subarea,
        "subarea": record.personnel_subarea,
        "position": record.position,
        "role": record.position,
        "section": record.section,
        "sub_section": record.sub_section,
        "ps_group": record.ps_group,
        "age": record.age,
        "flag_type": record.flag_type,
        "flagged_on": record.flagged_on.isoformat() if record.flagged_on else None,
        "source_file": record.source_file,
        "resolved": record.resolved,
        "job_run_id": record.job_run_id,
        "status_category": "new_hires",
        "date_done": None,
        "next_due": None,
    }


def _get_headers(ws, required_headers):
    headers = {}
    for col in range(1, ws.max_column + 1):
        headers[ws.cell(row=1, column=col).value] = col

    next_col = ws.max_column + 1
    for header in required_headers:
        if header not in headers:
            ws.cell(row=1, column=next_col, value=header)
            headers[header] = next_col
            next_col += 1
    return headers


def _find_sheet_row(ws, headers, employee_number, personnel_name, id_header, name_header):
    target_id = _normalise_lookup(employee_number)
    target_name = _normalise_lookup(personnel_name)
    for row_idx in range(2, ws.max_row + 1):
        row_id = _normalise_lookup(ws.cell(row=row_idx, column=headers[id_header]).value) if id_header in headers else ""
        row_name = _normalise_lookup(ws.cell(row=row_idx, column=headers[name_header]).value) if name_header in headers else ""
        if target_id and row_id == target_id:
            return row_idx
        if target_name and row_name == target_name:
            return row_idx
    return None


def _update_periodic_workbook(record, original_lookup, changes):
    from config import CONFIG

    wb = load_workbook(CONFIG["EXCEL_PATH"])
    ws = wb[CONFIG["PERIODICS_SHEET"]]
    headers = _get_headers(ws, PERIODIC_FIELD_TO_HEADER.values())
    row_idx = _find_sheet_row(ws, headers, original_lookup["employee_number"], original_lookup["personnel_name"], "EmployeeID", "Personnel Names")
    if not row_idx:
        raise ValueError("Could not locate employee in DCC PERIODICS sheet")

    for field, header in PERIODIC_FIELD_TO_HEADER.items():
        if field in changes:
          ws.cell(row=row_idx, column=headers[header], value=_coerce_workbook_value(field, getattr(record, field)))
    wb.save(CONFIG["EXCEL_PATH"])


def _update_new_hires_workbook(record, original_lookup, changes):
    from config import CONFIG

    wb = load_workbook(CONFIG["EXCEL_PATH"])
    ws = wb[CONFIG["NEW_EMP_SHEET"]]
    headers = _get_headers(ws, HEADCOUNT_FIELD_TO_HEADER.values())
    row_idx = _find_sheet_row(ws, headers, original_lookup["employee_number"], original_lookup["personnel_name"], "Pers.No.", "Personnel Names")
    if not row_idx:
        raise ValueError("Could not locate employee in New Employees sheet")

    for field, header in HEADCOUNT_FIELD_TO_HEADER.items():
        if field in changes:
            value = getattr(record, field)
            ws.cell(row=row_idx, column=headers[header], value=_coerce_workbook_value(field, value))
    wb.save(CONFIG["EXCEL_PATH"])


def _log_record_change(source, record, changed_fields, reason):
    db.session.add(RecordChangeLog(
        record_source=source,
        record_id=record.id,
        employee_number=getattr(record, "employee_number", None),
        personnel_name=getattr(record, "personnel_name", "Unknown"),
        changed_fields=json.dumps(changed_fields),
        change_reason=reason,
    ))


def _load_periodics_records():
    from config import CONFIG

    df = pd.read_excel(CONFIG["EXCEL_PATH"], sheet_name=CONFIG["PERIODICS_SHEET"])
    import logging; logging.getLogger(__name__).info("PERIODICS columns: %s", df.columns.tolist())
    if "Personnel Names" in df.columns:
        df = df[df["Personnel Names"].notna() & (df["Personnel Names"].astype(str).str.strip() != "")]
    if "UpdateStatus" in df.columns:
        df = df[df["UpdateStatus"].fillna("").astype(str).str.strip().str.lower() != "exited"]

    today = datetime.today()
    warning_days = CONFIG.get("DAYS_WARNING", 30)
    df["_next_due_dt"] = pd.to_datetime(df.get("NextDue"), errors="coerce")
    df["_days_remaining"] = (df["_next_due_dt"] - today).dt.days

    dedup_cols = [c for c in ["Personnel Names", "Employee Number", "EmployeeID", "Pers.No."] if c in df.columns]
    if dedup_cols:
        df = df.drop_duplicates(subset=dedup_cols, keep="first")

    records = []
    for _, row in df.iterrows():
        days_remaining = None if pd.isna(row.get("_days_remaining")) else int(row["_days_remaining"])
        status_category = _periodic_status_category(days_remaining, warning_days)

        records.append({
            "employee_number": _coalesce_int(row, "Employee Number", "EmployeeID", "Pers.No."),
            "personnel_name":  _clean_text(row.get("Personnel Names")),
            "gender":          _clean_text(row.get("Gender")),
            "role":            _coalesce_text(row, "Position", "Role"),
            "department":      _clean_text(row.get("Department")),
            "subarea":         _coalesce_text(row, "Personnel Subarea", "Sub Area", "SubArea", "Personnel Sub Area"),
            "ps_group":        _coalesce_text(row, "PS Group", "PSGroup", "PS group"),
            "grade": _clean_text(row.get("Grade")),
            "base": _clean_text(row.get("Base")),
            "date_done": _format_date_value(row.get("DateDone")),
            "next_due": _format_date_value(row.get("NextDue")),
            "update_status": _clean_text(row.get("UpdateStatus")),
            "update_date": _format_date_value(row.get("UpdateDate")),
            "days_remaining": days_remaining,
            "status_flag": _clean_text(row.get("StatusFlag")),
            "fallback_used": _clean_text(row.get("FallbackUsed")),
            "source_file": _clean_text(row.get("SourceFile")),
            "form_type": _clean_text(row.get("FormType")),
            "surname": _clean_text(row.get("Surname")),
            "clinic": _clean_text(row.get("Clinic")),
            "needs_review": _clean_text(row.get("NeedsReview")),
            "review_reason": _clean_text(row.get("ReviewReason")),
            "status_category": status_category,
        })
    return records


def _sync_periodics_to_db(records):
    # Remove duplicate DB rows — match by employee_number when present, else by name
    seen_names = {}
    for r in PeriodicRecord.query.order_by(PeriodicRecord.id).all():
        key = r.employee_number if r.employee_number is not None else (r.personnel_name or "")
        if key in seen_names:
            db.session.delete(r)
        else:
            seen_names[key] = r
    db.session.flush()

    # Build lookup: prefer employee_number key, fall back to name
    existing = {}
    for r in PeriodicRecord.query.all():
        if r.employee_number is not None:
            existing[r.employee_number] = r
        else:
            existing[r.personnel_name or ""] = r

    seen = set()
    for item in records:
        key = item["employee_number"] if item["employee_number"] is not None else (item["personnel_name"] or "")
        seen.add(key)
        record = existing.get(key)
        if not record:
            record = PeriodicRecord(
                employee_number=item["employee_number"],
                personnel_name=item["personnel_name"] or "Unknown",
            )
            db.session.add(record)
            existing[key] = record
        record.gender = item["gender"]
        record.role = item["role"]
        record.department = item["department"]
        record.subarea = item["subarea"]
        record.ps_group = item["ps_group"]
        record.grade = item["grade"]
        record.base = item["base"]
        record.date_done = item["date_done"]
        record.next_due = item["next_due"]
        record.update_status = item["update_status"]
        record.update_date = item["update_date"]
        record.days_remaining = item["days_remaining"]
        record.status_flag = item["status_flag"]
        record.fallback_used = item["fallback_used"]
        record.source_file = item["source_file"]
        record.form_type = item["form_type"]
        record.surname = item["surname"]
        record.clinic = item["clinic"]
        record.needs_review = item["needs_review"]
        record.review_reason = item["review_reason"]
        record.synced_at = datetime.utcnow()

    for key, record in existing.items():
        if key not in seen:
            db.session.delete(record)

    db.session.commit()


def _distinct_values(records, field_name):
    values = set()
    for record in records:
        value = getattr(record, field_name, None)
        if value is None:
            continue
        text = str(value).strip()
        if text:
            values.add(text)
    return sorted(values)


@api_bp.route("/department-breakdown")
@login_required
def department_breakdown():
    warning_days = current_app.config.get("DAYS_WARNING", 30)
    records = PeriodicRecord.query.all()

    dept_map = {}
    for r in records:
        dept = (r.department or "Unknown").strip()
        if dept not in dept_map:
            dept_map[dept] = {"name": dept, "overdue": 0, "due_soon": 0, "up_to_date": 0, "total": 0}
        cat = _periodic_status_category(r.days_remaining, warning_days)
        if cat in ("overdue", "due_soon", "up_to_date"):
            dept_map[dept][cat] += 1
            dept_map[dept]["total"] += 1

    # Sort by lowest compliance rate (most overdue+due_soon first), top 8 only
    def non_compliance_rate(d):
        if d["total"] == 0:
            return 0
        return (d["overdue"] + d["due_soon"]) / d["total"]

    result = sorted(dept_map.values(), key=non_compliance_rate, reverse=True)[:8]
    return jsonify(result)


@api_bp.route("/metrics")
@login_required
def metrics():
    from modules.exam_status import get_exam_status_counts
    total        = JobRun.query.count()
    done         = JobRun.query.filter_by(status="done").count()
    failed       = JobRun.query.filter_by(status="failed").count()
    conflicts    = ConflictLog.query.filter_by(resolved=False).count()
    hc_flags     = HeadcountRecord.query.filter_by(resolved=False).count()
    medic_flags  = MedicRecord.query.filter_by(resolved=False).count()
    pending_hires = HeadcountRecord.query.filter_by(flag_type="new", resolved=False).count()
    exam         = get_exam_status_counts()
    return jsonify({
        "total_runs":     total,
        "successful":     done,
        "failed":         failed,
        "conflicts":      conflicts,
        "hc_flags":       hc_flags,
        "medic_flags":    medic_flags,
        "overdue":        exam["overdue"],
        "due_soon":       exam["due_soon"],
        "up_to_date":     exam["up_to_date"],
        "pending_hires":  pending_hires,
    })


@api_bp.route("/jobs")
@login_required
def jobs():
    limit = min(int(request.args.get("limit", 50)), 500)
    runs  = JobRun.query.order_by(JobRun.started_at.desc()).limit(limit).all()
    return jsonify([r.to_dict() for r in runs])


@api_bp.route("/conflicts")
@login_required
def conflicts():
    include_resolved = request.args.get("resolved", "false").lower() == "true"
    query = ConflictLog.query
    if not include_resolved:
        query = query.filter_by(resolved=False)
    logs = query.order_by(ConflictLog.logged_at.desc()).all()
    return jsonify([c.to_dict() for c in logs])


@api_bp.route("/conflicts/<int:conflict_id>/resolve", methods=["PATCH"])
@login_required
@write_required
def resolve_conflict(conflict_id):
    conflict = ConflictLog.query.get(conflict_id)
    if not conflict:
        return jsonify({"error": "Conflict not found"}), 404
    conflict.resolved = True
    db.session.commit()
    return jsonify(conflict.to_dict())


@api_bp.route("/headcount")
@login_required
def headcount():
    flags = HeadcountRecord.query.filter_by(resolved=False)\
                .order_by(HeadcountRecord.flagged_on.desc()).all()
    return jsonify([h.to_dict() for h in flags])


@api_bp.route("/medic")
@login_required
def medic():
    records = MedicRecord.query.filter_by(resolved=False)\
                  .order_by(MedicRecord.flagged_on.desc()).all()
    return jsonify([m.to_dict() for m in records])


@api_bp.route("/run/<task>", methods=["POST"])
@login_required
@write_required
@limiter.limit("10 per minute")
def run_task(task):
    if task not in TASK_MAP and task != "full_cycle":
        return jsonify({"error": f"Unknown task: {task}"}), 400

    if not _task_lock.acquire(blocking=False):
        return jsonify({"error": "A task is already running. Please wait for it to finish."}), 429

    from app import create_app
    app = current_app._get_current_object()

    payload   = request.get_json(silent=True) or {}
    file_path = payload.get("file_path")

    def _run():
        try:
            with app.app_context():
                if task == "full_cycle":
                    import app.services.email_processor as ep
                    import app.services.booking_alerts as ba
                    ep.run()
                    ba.run()
                else:
                    import importlib
                    mod = importlib.import_module(TASK_MAP[task])
                    if file_path:
                        mod.run(file_path=file_path)
                    else:
                        mod.run()
        finally:
            _task_lock.release()

    thread = threading.Thread(target=_run, daemon=True)
    thread.start()
    return jsonify({"status": "queued", "task": task})

@api_bp.route("/biometric/stats")
@login_required
@biometrics_required
def biometric_stats():
    from config import CONFIG
    from datetime import timedelta

    result = {
        "total_employees": 0,
        "total_screened": 0,
        "male_count": 0,
        "female_count": 0,
        "avg_age": None,
        "recent_30_days": 0,
        "needs_review_count": 0,
        "coverage_pct": 0,
        "error": None,
    }

    result["total_employees"] = PeriodicRecord.query.count()
    result["needs_review_count"] = JobRun.query.filter_by(
        task_name="Biometric Screening", status="warning"
    ).count()

    bio_path = CONFIG.get("BIOMETRIC_EXCEL_PATH", "")
    bio_sheet = CONFIG.get("BIOMETRIC_SHEET", "Sheet1")

    if not bio_path or not os.path.isfile(bio_path):
        result["error"] = "Biometric data file not accessible"
        return jsonify(result)

    try:
        df = pd.read_excel(bio_path, sheet_name=bio_sheet)

        name_cols = [c for c in ["First name", "Last name"] if c in df.columns]
        if name_cols:
            df = df[df[name_cols].notna().any(axis=1)]

        result["total_screened"] = len(df)

        if "Gender" in df.columns:
            gender_norm = df["Gender"].astype(str).str.strip().str.upper()
            result["male_count"] = int(gender_norm.isin(["M", "MALE"]).sum())
            result["female_count"] = int(gender_norm.isin(["F", "FEMALE"]).sum())

        age_col = next((c for c in df.columns if "age" in c.lower()), None)
        if age_col:
            ages = pd.to_numeric(df[age_col], errors="coerce").dropna()
            if len(ages):
                result["avg_age"] = round(float(ages.mean()), 1)

        if "Date" in df.columns:
            dates = pd.to_datetime(df["Date"], dayfirst=True, errors="coerce")
            cutoff = pd.Timestamp.now() - pd.Timedelta(days=30)
            result["recent_30_days"] = int((dates >= cutoff).sum())

        total = result["total_employees"]
        if total > 0:
            result["coverage_pct"] = round(
                min(result["total_screened"] / total * 100, 100), 1
            )

    except Exception as e:
        result["error"] = str(e)

    return jsonify(result)


@api_bp.route("/biometric/extract", methods=["POST"])
@login_required
@biometrics_required
@limiter.limit("10 per minute")
@write_required
def biometric_extract():
    """Phase 1: extract from PDF, return data for review modal."""
    payload   = request.get_json(silent=True) or {}
    file_path = payload.get("file_path")
    if not file_path:
        return jsonify({"error": "file_path is required"}), 400
    try:
        from app.services.biometric_processor import extract
        data = extract(file_path)
        # Don't send raw text to the browser — too large
        data.pop("_raw_text", None)
        return jsonify(data)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@api_bp.route("/biometric/confirm", methods=["POST"])
@login_required
@biometrics_required
@limiter.limit("10 per minute")
@write_required
def biometric_confirm():
    """Phase 2: write reviewed/edited data to Excel, log JobRun."""
    payload   = request.get_json(silent=True) or {}
    data      = payload.get("data", {})
    file_path = payload.get("file_path", "")
    if not data:
        return jsonify({"error": "data is required"}), 400
    try:
        from app.services.biometric_processor import confirm
        job = confirm(data, file_path)
        return jsonify(job)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@api_bp.route("/biometric/records")
@login_required
@biometrics_required
def biometric_records():
    from config import CONFIG
    bio_path  = CONFIG.get("BIOMETRIC_EXCEL_PATH", "")
    bio_sheet = CONFIG.get("BIOMETRIC_SHEET", "Sheet1")

    if not bio_path or not os.path.isfile(bio_path):
        return jsonify({"error": "Biometric data file not accessible"})

    try:
        df = pd.read_excel(bio_path, sheet_name=bio_sheet, dtype=str)
        df = df.fillna("")
        name_cols = [c for c in ["First name", "Last name"] if c in df.columns]
        if name_cols:
            mask = df[name_cols].apply(lambda col: col.str.strip() != "").any(axis=1)
            df = df[mask]
        records = []
        for orig_idx, row in df.iterrows():
            rec = {k: str(v).strip() for k, v in row.items()}
            rec["row_idx"] = int(orig_idx)
            records.append(rec)
        return jsonify(records)
    except Exception as e:
        return jsonify({"error": str(e)})


@api_bp.route("/biometric/records/<int:row_idx>", methods=["PATCH"])
@login_required
@biometrics_required
@write_required
def biometric_record_update(row_idx):
    from config import CONFIG
    payload   = request.get_json(silent=True) or {}
    changes   = payload.get("changes", {})
    bio_path  = CONFIG.get("BIOMETRIC_EXCEL_PATH", "")
    bio_sheet = CONFIG.get("BIOMETRIC_SHEET", "Sheet1")

    if not bio_path or not os.path.isfile(bio_path):
        return jsonify({"error": "Biometric data file not accessible"}), 400

    try:
        wb = load_workbook(bio_path)
        ws = wb[bio_sheet]
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        col_index  = {str(c).strip(): i + 1 for i, c in enumerate(header_row) if c is not None}
        excel_row  = row_idx + 2   # pandas 0-based index → Excel row (header is row 1)
        for key, value in changes.items():
            col_idx = col_index.get(key)
            if col_idx:
                ws.cell(row=excel_row, column=col_idx, value=value or None)
        wb.save(bio_path)
        return jsonify({"ok": True})
    except PermissionError:
        return jsonify({"error": "File is open in another program. Close it and try again."}), 400
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@api_bp.route("/task-summary")
@login_required
def task_summary():
    runs    = JobRun.query.all()
    summary = {}
    for run in runs:
        t = run.task_name
        if t not in summary:
            summary[t] = {"task": t, "total": 0, "done": 0, "failed": 0}
        summary[t]["total"] += 1
        if run.status == "done":
            summary[t]["done"] += 1
        elif run.status == "failed":
            summary[t]["failed"] += 1
    result = []
    for s in summary.values():
        s["success_pct"] = round(s["done"] / s["total"] * 100) if s["total"] else 0
        result.append(s)
    return jsonify(sorted(result, key=lambda x: x["task"]))


@api_bp.route("/new-hires")
@login_required
def new_hires():
    records = HeadcountRecord.query\
        .filter_by(flag_type="new", resolved=False)\
        .order_by(HeadcountRecord.flagged_on.desc()).limit(20).all()
    return jsonify([r.to_dict() for r in records])


@api_bp.route("/employee-status")
@login_required
def employee_status():
    category = request.args.get("category", "overdue")
    if category not in STATUS_FILTERS:
        return jsonify({"error": "Unknown category"}), 400

    if category == "new_hires":
        records = HeadcountRecord.query\
            .filter_by(flag_type="new", resolved=False)\
            .order_by(HeadcountRecord.flagged_on.desc()).all()
        payload = [_serialise_headcount_record(r) for r in records]
        return jsonify(payload)

    try:
        records = _load_periodics_records()
        _sync_periodics_to_db(records)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

    payload = [_serialise_periodic_record(r) for r in PeriodicRecord.query.all() if _periodic_status_category(r.days_remaining, current_app.config.get("DAYS_WARNING", 30)) == category]
    payload.sort(key=lambda r: ((r.get("days_remaining") is None), r.get("days_remaining", 999999), r.get("personnel_name") or ""))
    return jsonify(payload)


@api_bp.route("/employee-status-options")
@login_required
def employee_status_options():
    try:
        records = _load_periodics_records()
        _sync_periodics_to_db(records)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

    periodic_records = PeriodicRecord.query.all()
    headcount_records = HeadcountRecord.query.all()

    return jsonify({
        "periodic": {
            "department": _distinct_values(periodic_records, "department"),
            "subarea": _distinct_values(periodic_records, "subarea"),
            "ps_group": _distinct_values(periodic_records, "ps_group"),
            "clinic": _distinct_values(periodic_records, "clinic"),
            "base": _distinct_values(periodic_records, "base"),
            "grade": _distinct_values(periodic_records, "grade"),
            "update_status": _distinct_values(periodic_records, "update_status"),
            "status_flag": _distinct_values(periodic_records, "status_flag"),
        },
        "headcount": {
            "department": _distinct_values(headcount_records, "department"),
            "personnel_subarea": _distinct_values(headcount_records, "personnel_subarea"),
            "ps_group": _distinct_values(headcount_records, "ps_group"),
            "section": _distinct_values(headcount_records, "section"),
            "sub_section": _distinct_values(headcount_records, "sub_section"),
            "position": _distinct_values(headcount_records, "position"),
            "flag_type": _distinct_values(headcount_records, "flag_type"),
        },
    })


@api_bp.route("/employee-status/<source>/<int:record_id>", methods=["PATCH"])
@login_required
@write_required
@limiter.limit("30 per minute")
def update_employee_status_record(source, record_id):
    payload = request.get_json(silent=True) or {}
    changes = payload.get("changes", {})
    reason = (payload.get("reason") or "").strip()

    if not reason:
        return jsonify({"error": "A reason for the change is required"}), 400

    if source == "periodic":
        record = PeriodicRecord.query.filter_by(id=record_id).first()
        if not record:
            return jsonify({"error": "Record not found"}), 404
        field_map = PERIODIC_FIELD_TO_HEADER
        original_lookup = {
            "employee_number": record.employee_number,
            "personnel_name": record.personnel_name,
        }
        changed_fields = {}

        try:
            for field in field_map:
                if field not in changes:
                    continue
                new_value = _coerce_value(field, changes.get(field))
                old_value = getattr(record, field)
                if old_value != new_value:
                    changed_fields[field] = {"from": old_value, "to": new_value}
                    setattr(record, field, new_value)

            if not changed_fields:
                return jsonify({"record": _serialise_periodic_record(record), "message": "No changes detected"})

            record.days_remaining = None
            if record.next_due:
                parsed_next_due = pd.to_datetime(record.next_due, errors="coerce")
                if not pd.isna(parsed_next_due):
                    record.days_remaining = int((parsed_next_due.to_pydatetime() - datetime.today()).days)
            record.synced_at = datetime.utcnow()
            _update_periodic_workbook(record, original_lookup, changed_fields)
            _log_record_change("periodic", record, changed_fields, reason)
            db.session.commit()
            return jsonify({"record": _serialise_periodic_record(record)})
        except Exception as e:
            db.session.rollback()
            return jsonify({"error": str(e)}), 500

    if source == "headcount":
        record = HeadcountRecord.query.filter_by(id=record_id).first()
        if not record:
            return jsonify({"error": "Record not found"}), 404
        original_lookup = {
            "employee_number": record.employee_number,
            "personnel_name": record.personnel_name,
        }
        changed_fields = {}

        try:
            for field in HEADCOUNT_FIELD_TO_HEADER:
                if field not in changes:
                    continue
                new_value = _coerce_value(field, changes.get(field))
                old_value = getattr(record, field)
                if old_value != new_value:
                    changed_fields[field] = {"from": old_value.isoformat() if hasattr(old_value, "isoformat") and old_value else old_value, "to": new_value.isoformat() if hasattr(new_value, "isoformat") and new_value else new_value}
                    setattr(record, field, new_value)

            if not changed_fields:
                return jsonify({"record": _serialise_headcount_record(record), "message": "No changes detected"})

            _update_new_hires_workbook(record, original_lookup, changed_fields)
            _log_record_change("headcount", record, changed_fields, reason)
            db.session.commit()
            return jsonify({"record": _serialise_headcount_record(record)})
        except Exception as e:
            db.session.rollback()
            return jsonify({"error": str(e)}), 500

    return jsonify({"error": "Unsupported record source"}), 400


@api_bp.route("/record-change-logs")
@login_required
def record_change_logs():
    source = request.args.get("source", "all")
    limit = min(int(request.args.get("limit", 100)), 500)
    query = RecordChangeLog.query.order_by(RecordChangeLog.changed_at.desc())

    if source in {"periodic", "headcount"}:
        query = query.filter_by(record_source=source)
    elif source != "all":
        return jsonify({"error": "Unknown source filter"}), 400

    logs = query.limit(limit).all()
    payload = []
    for log in logs:
        item = log.to_dict()
        try:
            item["changed_fields"] = json.loads(item["changed_fields"])
        except Exception:
            item["changed_fields"] = {}
        payload.append(item)
    return jsonify(payload)


@api_bp.route("/files/<folder>")
@login_required
def list_files(folder):
    from config import CONFIG
    folders = {
        "headcount": CONFIG.get("HC_FOLDER", ""),
        "pdf":       CONFIG.get("SAVE_FOLDER", ""),
        "medic":     CONFIG.get("MEDIC_ID_FOLDER", ""),
        "biometric": CONFIG.get("BIOMETRIC_PDF_FOLDER", ""),
    }
    if folder not in folders:
        return jsonify({"error": "Unknown folder"}), 400

    path = folders[folder]
    if not path:
        return jsonify({"error": f"No path configured for '{folder}'"}), 500

    try:
        files = [
            f for f in os.listdir(path)
            if f.lower().endswith((".xlsx", ".xls", ".csv", ".pdf"))
        ]
        files.sort(reverse=True)
        return jsonify({"folder": path, "files": files})
    except Exception as e:
        return jsonify({"error": str(e)}), 500
