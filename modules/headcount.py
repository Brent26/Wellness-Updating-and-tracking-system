"""
modules/headcount.py
Reconciles a monthly headcount file against the Periodics sheet.

Matching strategy
─────────────────
1. Exact normalised-name match (lower-case, collapsed whitespace).
2. Fuzzy match (token_set_ratio >= FUZZY_THRESHOLD) to handle middle-name
   differences and minor spelling variations — e.g. "Nametso Chephethe"
   correctly resolves to "Nametso Audrey Chephethe" and neither is flagged.
3. Exits-sheet guard — names already recorded as leavers are never
   re-flagged as new hires, even if they reappear in the headcount file.

Applied in BOTH directions:
  • Suppresses false "new hire" flags  (name in HC but not in Periodics)
  • Suppresses false "exit" flags      (name in Periodics but not in HC)
"""

import re
import os
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config import CONFIG

try:
    from thefuzz import fuzz as _fuzz
    _FUZZY_OK = True
except ImportError:
    _FUZZY_OK = False

# Minimum token_set_ratio score to treat two names as the same person.
# At 90: missing middle names score 100 (always caught).
#        Minor typos score ~92 (caught).
#        Different surnames score ~67 (never matched).
FUZZY_THRESHOLD = 90


# ── Helpers ───────────────────────────────────────────────────────────────────

def _normalise(name):
    if pd.isna(name):
        return ""
    return re.sub(r"\s+", " ", str(name).strip().lower())


def _has_fuzzy_match(name, name_set, threshold=FUZZY_THRESHOLD):
    """
    Return True if *name* scores >= threshold against any entry in *name_set*.
    Uses token_set_ratio so that a subset of tokens (missing middle name)
    scores 100 against the superset.
    Falls back to False when thefuzz is not installed.
    """
    if not _FUZZY_OK or not name or not name_set:
        return False
    for candidate in name_set:
        if candidate and _fuzz.token_set_ratio(name, candidate) >= threshold:
            return True
    return False


def _thin_border():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)


def _style_header(cell, bg="1F4E79"):
    cell.font      = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    cell.fill      = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center",
                               wrap_text=True)
    cell.border    = _thin_border()


def _style_cell(cell, bg=None):
    cell.font      = Font(name="Arial", size=10)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border    = _thin_border()
    if bg:
        cell.fill = PatternFill("solid", start_color=bg)


# ── Reconciler ────────────────────────────────────────────────────────────────

class HeadcountReconciler:

    def __init__(self, ui=None):
        self.ui          = ui
        self.path        = CONFIG["EXCEL_PATH"]
        self.sheet       = CONFIG["PERIODICS_SHEET"]
        self.new_sheet   = CONFIG["NEW_EMP_SHEET"]
        self.exits_sheet = CONFIG["EXITS_SHEET"]

    # ── Public ────────────────────────────────────────────────────

    def reconcile(self, hc_path):
        if self.ui:
            self.ui.info(f"Reconciling: {os.path.basename(hc_path)}")

        # ── 1. Load incoming headcount file ───────────────────────
        try:
            hc_df = (pd.read_csv(hc_path)
                     if hc_path.endswith(".csv")
                     else pd.read_excel(hc_path))
        except Exception as e:
            if self.ui: self.ui.error(f"Cannot read headcount: {e}")
            return pd.DataFrame(), pd.DataFrame()

        if "Personnel Names" not in hc_df.columns:
            if self.ui:
                self.ui.warn("'Personnel Names' column missing — skipping")
            return pd.DataFrame(), pd.DataFrame()

        hc_df["_norm"] = hc_df["Personnel Names"].apply(_normalise)

        # ── 2. Load Periodics sheet ────────────────────────────────
        try:
            per_df = pd.read_excel(self.path, sheet_name=self.sheet)
        except Exception as e:
            if self.ui: self.ui.error(f"Cannot read Periodics sheet: {e}")
            return pd.DataFrame(), pd.DataFrame()

        if "Personnel Names" not in per_df.columns:
            if self.ui:
                self.ui.warn("'Personnel Names' missing in Periodics — skipping")
            return pd.DataFrame(), pd.DataFrame()

        per_df["_norm"] = per_df["Personnel Names"].apply(_normalise)

        # ── 3. Load Exits sheet — known leavers guard ─────────────
        # Anyone already recorded as a leaver must never be re-flagged as a
        # new hire, even if their name reappears in a later headcount file.
        exits_names = set()
        try:
            exits_df = pd.read_excel(self.path, sheet_name=self.exits_sheet)
            if "Personnel Names" in exits_df.columns:
                exits_names = {
                    n for n in exits_df["Personnel Names"].apply(_normalise)
                    if n
                }
                if self.ui and exits_names:
                    self.ui.info(
                        f"Exits sheet: {len(exits_names)} known leaver(s) loaded — "
                        f"will not be re-flagged as new hires"
                    )
        except Exception as e:
            if self.ui:
                self.ui.warn(
                    f"Exits sheet ('{self.exits_sheet}') could not be read — "
                    f"exits guard skipped: {e}"
                )

        hc_names  = set(hc_df["_norm"]) - {""}
        per_names = set(per_df["_norm"]) - {""}

        # ── 4. Exact set differences ───────────────────────────────
        new_exact    = hc_names  - per_names   # in HC, not in Periodics
        exited_exact = per_names - hc_names    # in Periodics, not in HC

        # ── 5. Fuzzy refinement — both directions ──────────────────
        # "New" candidate suppressed if it fuzzy-matches an existing periodic name
        # (handles missing middle names, minor typos, etc.)
        new_names = {
            n for n in new_exact
            if not _has_fuzzy_match(n, per_names)
        }

        # "Exited" candidate suppressed if it fuzzy-matches a headcount name
        # (prevents flagging "Nametso Audrey Chephethe" as exited when the HC
        # only has "Nametso Chephethe" — same person, different name format)
        exited_names = {
            n for n in exited_exact
            if not _has_fuzzy_match(n, hc_names)
        }

        # ── 6. Exits-sheet guard ───────────────────────────────────
        # A name that passed the fuzzy filter but is already on the exits list
        # is a known leaver reappearing in the headcount — do not flag as new.
        new_names = {
            n for n in new_names
            if n not in exits_names
            and not _has_fuzzy_match(n, exits_names)
        }

        if self.ui:
            self.ui.result("New employees",    len(new_names))
            self.ui.result("Possible leavers", len(exited_names))

        today_str = datetime.today().strftime("%d-%b-%Y")

        # ── 7. Flag leavers in Periodics sheet ────────────────────
        if exited_names:
            for col in ["UpdateStatus", "UpdateDate"]:
                if col not in per_df.columns:
                    per_df[col] = None
                per_df[col] = per_df[col].astype(object)
            mask = per_df["_norm"].isin(exited_names)
            per_df.loc[mask, "UpdateStatus"] = "Exited"
            per_df.loc[mask, "UpdateDate"]   = today_str
            clean = per_df.drop(columns=["_norm"])
            with pd.ExcelWriter(self.path, engine="openpyxl", mode="a",
                                if_sheet_exists="replace") as writer:
                clean.to_excel(writer, sheet_name=self.sheet, index=False)
            if self.ui:
                self.ui.warn(f"Flagged {len(exited_names)} employee(s) as Exited")

        # ── 8. Write new employees sheet ──────────────────────────
        new_rows = hc_df[hc_df["_norm"].isin(new_names)].copy()
        new_rows = new_rows.drop(columns=["_norm"])
        new_rows["Flagged On"]   = today_str
        new_rows["Action Taken"] = "Pending"
        self._write_new_emp_sheet(new_rows)

        if self.ui:
            self.ui.success(
                f"{len(new_names)} new employee(s) written to "
                f"'{self.new_sheet}' sheet"
            )

        exited_rows = (
            per_df[per_df["_norm"].isin(exited_names)]
            .drop(columns=["_norm"])
            .copy()
        )
        return new_rows, exited_rows

    # ── Private ───────────────────────────────────────────────────

    def _write_new_emp_sheet(self, df):
        wb = load_workbook(self.path)
        if self.new_sheet in wb.sheetnames:
            del wb[self.new_sheet]
        ws   = wb.create_sheet(self.new_sheet)
        cols = list(df.columns)

        for c_idx, col in enumerate(cols, 1):
            _style_header(ws.cell(row=1, column=c_idx, value=col))
        ws.row_dimensions[1].height = 28

        for r_idx, (_, row) in enumerate(df.iterrows(), 2):
            for c_idx, col in enumerate(cols, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=row[col])
                _style_cell(cell, bg="FFF2CC")

        for c_idx, col in enumerate(cols, 1):
            vals = [str(col)] + [str(df.iloc[r][col]) for r in range(len(df))]
            ws.column_dimensions[get_column_letter(c_idx)].width = \
                min(max(len(v) for v in vals) + 4, 40)

        ws.freeze_panes = "A2"
        wb.save(self.path)
