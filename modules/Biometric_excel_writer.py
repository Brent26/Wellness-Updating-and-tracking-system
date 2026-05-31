"""
modules/Biometric_excel_writer.py

Appends a new biometric screening row to the target Excel sheet.
  Target file  : CONFIG["BIOMETRIC_EXCEL_PATH"]
  Target sheet : CONFIG["BIOMETRIC_SHEET"]  (default "Sheet1")

Header row is read dynamically — column order in the spreadsheet does not matter.
Only the columns listed in COLUMN_MAP are written; all other columns are left empty.
"""

import os
from datetime import datetime
from openpyxl import load_workbook
from config import CONFIG

# Maps extractor dict keys  →  exact column header names in the spreadsheet
COLUMN_MAP = {
    "Date":              "Date",
    "First name":        "First name",
    "Last name":         "Last name",
    "Age of employee":   "Age of employee",
    "Gender":            "Gender",
    "Glucose level":     "Glucose level",
    "Total Cholesterol": "Total Cholesterol",
    "HIV":               "HIV",
    "PSA":               "PSA",
    "Mammo":             "Mammo",
    "Pap Smear":         "PapSmear",
}


class BiometricExcelWriter:

    def __init__(self, ui=None):
        self._ui    = ui
        self._path  = CONFIG["BIOMETRIC_EXCEL_PATH"]
        self._sheet = CONFIG["BIOMETRIC_SHEET"]

    # ── Public ────────────────────────────────────────────────────

    def write(self, data: dict) -> bool:
        """
        Append *data* as a new row.  Returns True on success, False on failure.
        *data* keys match the extractor output (see COLUMN_MAP).
        """
        if not os.path.isfile(self._path):
            self._err(f"Excel file not found: {self._path}")
            return False

        try:
            wb = load_workbook(self._path)
        except Exception as e:
            self._err(f"Could not open workbook: {e}")
            return False

        if self._sheet not in wb.sheetnames:
            self._err(
                f"Sheet '{self._sheet}' not found. "
                f"Available: {', '.join(wb.sheetnames)}"
            )
            return False

        ws = wb[self._sheet]

        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if header_row is None:
            self._err("Sheet appears to be empty — no header row found.")
            return False

        col_index = {
            str(cell).strip(): idx + 1
            for idx, cell in enumerate(header_row)
            if cell is not None
        }

        next_row = ws.max_row + 1

        for data_key, sheet_col in COLUMN_MAP.items():
            value = data.get(data_key)
            if value is None:
                continue
            idx = col_index.get(sheet_col)
            if idx is None:
                self._warn(f"Column '{sheet_col}' not found in header — skipped.")
                continue
            if isinstance(value, datetime):
                value = value.strftime("%d/%m/%Y")
            ws.cell(row=next_row, column=idx, value=value)

        try:
            wb.save(self._path)
        except PermissionError:
            self._err(
                "Could not save — Excel file may be open in another program. "
                "Please close it and try again."
            )
            return False
        except Exception as e:
            self._err(f"Could not save workbook: {e}")
            return False

        return True

    # ── Helpers ───────────────────────────────────────────────────

    def _err(self, msg):
        if self._ui:
            self._ui.error(msg)
        else:
            print(f"[ERROR] {msg}")

    def _warn(self, msg):
        if self._ui:
            self._ui.warn(msg)
        else:
            print(f"[WARN] {msg}")
