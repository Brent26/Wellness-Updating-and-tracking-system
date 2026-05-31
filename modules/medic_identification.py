"""
modules/medic_identification.py

Reconciles a MEDIC IDENTIFICATION file against DCC PERIODICS and EXITS sheets.

MEDIC IDENTIFICATION columns (header always row 1):
    Date | Personnel Names | Type of Medical

Three reconciliation tasks:

  1. DATE COMPARISON
     Match people by name across both files.
     If DateDone in DCC PERIODICS differs from Date in MEDIC IDENTIFICATION,
     write the record to the output file.

  2. EXIT DETECTION (not yet in EXITS sheet)
     If Type of Medical = "Exit" and the person IS in DCC PERIODICS
     but NOT yet in EXITS sheet — write their full record to the
     output file under the "Unprocessed Exits" tab.

  3. EXIT DATE SYNC (already in EXITS sheet)
     If Type of Medical = "Exit" and the person IS already in the
     EXITS sheet — compare the Date in MEDIC IDENTIFICATION against
     the Exit Date in EXITS sheet. If they differ, update the EXITS
     sheet with the MEDIC IDENTIFICATION date.
"""

import os
import datetime as _dt
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config import CONFIG
from modules.name_utils import normalise as _norm_fn, find_name_in_df

# ── Styling constants ─────────────────────────────────────────
_TEAL       = "1F6B75"
_TEAL_LIGHT = "E8F4F5"
_AMBER      = "B8860B"
_AMBER_BG   = "FFF3CD"
_RED        = "C00000"
_RED_BG     = "FCE4D6"
_GREEN      = "1F6B45"
_GREEN_BG   = "E6FFE6"
_WHITE      = "FFFFFF"
_GREY_BG    = "F2F2F2"


def _thin():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)


def _hdr(cell, bg=_TEAL):
    cell.font      = Font(bold=True, color=_WHITE, name="Arial", size=10)
    cell.fill      = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center",
                               wrap_text=True)
    cell.border    = _thin()


def _dat(cell, bg=None, bold=False, color="000000"):
    cell.font      = Font(name="Arial", size=10, bold=bold, color=color)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border    = _thin()
    if bg:
        cell.fill = PatternFill("solid", start_color=bg)


def _autowidth(ws, cols, extra=4, max_w=45):
    for i, col in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(i)].width = \
            min(len(str(col)) + extra, max_w)


class MedicIdentificationReconciler:

    _MISMATCH_COLS = [
        "Personnel Names",
        "Employee Number",
        "Department",
        "Position",
        "PS Group",
        "Personnel Subarea",
        "Date in DCC PERIODICS",
        "Date in MEDIC IDENTIFICATION",
        "Date to Update To",
        "Difference (days)",
        "Type of Medical",
        "Flagged On",
    ]

    _EXIT_COLS = [
        "Personnel Names",
        "Employee Number",
        "Department",
        "Personnel Subarea",
        "Position",
        "Section",
        "Sub Section",
        "Age",
        "PS Group",
        "Last Medical (DCC PERIODICS)",
        "Date in MEDIC IDENTIFICATION",
        "Type of Medical",
        "Exit Status",
        "Flagged On",
    ]

    def __init__(self, ui=None):
        self.ui          = ui
        self.excel_path  = CONFIG["EXCEL_PATH"]
        self.per_sheet   = CONFIG["PERIODICS_SHEET"]
        self.exits_sheet = CONFIG["EXITS_SHEET"]
        self.out_folder  = CONFIG["MEDIC_ID_OUTPUT_FOLDER"]

    def reconcile(self, medic_path):
        if self.ui:
            self.ui.info(f"Loading: {os.path.basename(medic_path)}")

        # ── Load MEDIC IDENTIFICATION ─────────────────────────
        try:
            xl_file      = pd.ExcelFile(medic_path)
            year_sheet   = str(_dt.datetime.today().year)
            sheet_to_use = (
                year_sheet
                if year_sheet in xl_file.sheet_names
                else xl_file.sheet_names[0]
            )
            if self.ui:
                self.ui.info(f"Using sheet: '{sheet_to_use}'")

            raw        = pd.read_excel(medic_path, sheet_name=sheet_to_use, header=None)
            header_row = None

            for i, row in raw.iterrows():
                vals = [str(v).strip() for v in row.values]
                if "Date" in vals and "Type of Medical" in vals:
                    header_row = i
                    break

            if header_row is None:
                for i, row in raw.iterrows():
                    vals = [str(v).strip() for v in row.values]
                    if "Date" in vals and ("Name" in vals or "Surname" in vals):
                        header_row = i
                        break

            if header_row is None:
                if self.ui:
                    self.ui.error(
                        "Cannot locate header row in MEDIC IDENTIFICATION.\n"
                        "Expected a row containing 'Date' and 'Type of Medical'."
                    )
                return

            medic_df = pd.read_excel(
                medic_path, sheet_name=sheet_to_use, header=header_row
            )

        except Exception as e:
            if self.ui:
                self.ui.error(f"Cannot read MEDIC IDENTIFICATION: {e}")
            return

        # ── Auto-build Personnel Names if split ───────────────
        if "Personnel Names" not in medic_df.columns:
            if "Name" in medic_df.columns and "Surname" in medic_df.columns:
                medic_df["Personnel Names"] = (
                    medic_df["Name"].astype(str).str.strip()
                    + " "
                    + medic_df["Surname"].astype(str).str.strip()
                )
                if self.ui:
                    self.ui.info("Auto-combined 'Name' + 'Surname' → 'Personnel Names'")
            else:
                if self.ui:
                    self.ui.error(
                        f"Cannot find 'Personnel Names' column.\n"
                        f"Found: {list(medic_df.columns)}"
                    )
                return

        required = {"Date", "Personnel Names", "Type of Medical"}
        missing  = required - set(medic_df.columns)
        if missing:
            if self.ui:
                self.ui.error(f"Missing columns: {missing}")
            return

        medic_df["Date"]            = pd.to_datetime(medic_df["Date"], dayfirst=True, errors="coerce")
        medic_df["_norm"]           = medic_df["Personnel Names"].apply(self._norm)
        medic_df["Type of Medical"] = medic_df["Type of Medical"].astype(str).str.strip()

        # ── Load DCC PERIODICS ────────────────────────────────
        try:
            per_df = pd.read_excel(self.excel_path, sheet_name=self.per_sheet)
        except Exception as e:
            if self.ui:
                self.ui.error(f"Cannot read DCC PERIODICS: {e}")
            return

        per_df["DateDone"] = pd.to_datetime(per_df.get("DateDone", pd.Series()), errors="coerce")
        per_df["_norm"]    = per_df["Personnel Names"].apply(self._norm)

        # ── Load EXITS sheet ──────────────────────────────────
        exits_names = set()
        exits_df    = pd.DataFrame()
        try:
            wb_check = load_workbook(self.excel_path, read_only=True)
            if self.exits_sheet in wb_check.sheetnames:
                wb_check.close()
                exits_df = pd.read_excel(self.excel_path, sheet_name=self.exits_sheet)
                if "Personnel Names" in exits_df.columns:
                    exits_df["_norm"] = exits_df["Personnel Names"].apply(self._norm)
                    exits_names = set(exits_df["_norm"]) - {""}
            else:
                wb_check.close()
        except Exception as e:
            if self.ui:
                self.ui.warn(f"Could not load EXITS sheet: {e}")

        today_str = datetime.today().strftime("%d-%b-%Y")

        mismatches        = self._compare_dates(medic_df, per_df, today_str)
        unprocessed_exits = self._find_unprocessed_exits(medic_df, per_df, exits_names, today_str)
        synced            = self._sync_exit_dates(medic_df, exits_df, exits_names)

        if mismatches or unprocessed_exits:
            self._write_output(mismatches, unprocessed_exits, today_str)
        else:
            if self.ui:
                self.ui.success("No date mismatches or unprocessed exits found")

        if self.ui:
            self.ui.result("Date mismatches found",   len(mismatches))
            self.ui.result("Unprocessed exits found", len(unprocessed_exits))
            self.ui.result("Exit dates synced",       synced)

        return mismatches, unprocessed_exits, synced

    def _compare_dates(self, medic_df, per_df, today_str):
        mismatches = []
        for _, m_row in medic_df.iterrows():
            if not m_row["_norm"]:
                continue
            idx, match_type, score = find_name_in_df(m_row["Personnel Names"], per_df)
            if idx is None:
                continue
            p_row  = per_df.loc[idx]
            p_date = p_row.get("Last Medical") or p_row.get("DateDone")
            m_date = m_row["Date"]
            if pd.isna(p_date) or pd.isna(m_date):
                continue
            p_date_d = pd.Timestamp(p_date).normalize()
            m_date_d = pd.Timestamp(m_date).normalize()
            if p_date_d != m_date_d:
                diff = (m_date_d - p_date_d).days
                mismatches.append({
                    "Personnel Names":              m_row["Personnel Names"],
                    "Employee Number":              p_row.get("Employee Number", ""),
                    "Department":                   p_row.get("Department", ""),
                    "Position":                     p_row.get("Position", ""),
                    "PS Group":                     p_row.get("PS Group", ""),
                    "Personnel Subarea":            p_row.get("Personnel Subarea", ""),
                    "Date in DCC PERIODICS":        p_date_d.strftime("%d-%b-%Y"),
                    "Date in MEDIC IDENTIFICATION": m_date_d.strftime("%d-%b-%Y"),
                    "Date to Update To":            m_date_d.strftime("%d-%b-%Y"),
                    "Difference (days)":            diff,
                    "Type of Medical":              m_row.get("Type of Medical", ""),
                    "Flagged On":                   today_str,
                })
        return mismatches

    def _find_unprocessed_exits(self, medic_df, per_df, exits_names, today_str):
        unprocessed = []
        exit_records = medic_df[medic_df["Type of Medical"].str.lower() == "exit"]
        for _, m_row in exit_records.iterrows():
            if not m_row["_norm"]:
                continue
            from modules.name_utils import fuzzy_match
            exits_series = pd.Series(list(exits_names), index=range(len(exits_names)))
            exits_idx, exits_score = fuzzy_match(m_row["_norm"], exits_series)
            if exits_idx is not None:
                continue
            idx, match_type, score = find_name_in_df(m_row["Personnel Names"], per_df)
            if idx is None:
                continue
            p_row = per_df.loc[idx]
            last_med = p_row.get("Last Medical") or p_row.get("DateDone")
            unprocessed.append({
                "Personnel Names":              m_row["Personnel Names"],
                "Employee Number":              p_row.get("Employee Number", ""),
                "Department":                   p_row.get("Department", ""),
                "Personnel Subarea":            p_row.get("Personnel Subarea", ""),
                "Position":                     p_row.get("Position", ""),
                "Section":                      p_row.get("Section", ""),
                "Sub Section":                  p_row.get("Sub Section", ""),
                "Age":                          p_row.get("Age", ""),
                "PS Group":                     p_row.get("PS Group", ""),
                "Last Medical (DCC PERIODICS)": (
                    pd.Timestamp(last_med).strftime("%d-%b-%Y")
                    if pd.notna(last_med) else ""
                ),
                "Date in MEDIC IDENTIFICATION": (
                    m_row["Date"].strftime("%d-%b-%Y")
                    if pd.notna(m_row["Date"]) else ""
                ),
                "Type of Medical":              "Exit",
                "Exit Status":                  "In DCC PERIODICS — not yet moved to EXITS",
                "Flagged On":                   today_str,
            })
        return unprocessed

    def _sync_exit_dates(self, medic_df, exits_df, exits_names):
        if exits_df.empty or "_norm" not in exits_df.columns:
            return 0
        exit_records = medic_df[medic_df["Type of Medical"].str.lower() == "exit"]
        synced   = 0
        modified = False
        for _, m_row in exit_records.iterrows():
            if not m_row["_norm"]:
                continue
            m_date = m_row["Date"]
            if pd.isna(m_date):
                continue
            idx, match_type, score = find_name_in_df(m_row["Personnel Names"], exits_df)
            if idx is None:
                continue
            date_col = next(
                (c for c in exits_df.columns if "exit" in c.lower() and "date" in c.lower()),
                None
            )
            if date_col:
                exits_df[date_col] = exits_df[date_col].astype(object)
                existing_date = pd.to_datetime(exits_df.loc[idx, date_col], errors="coerce")
                if pd.isna(existing_date) or (
                    pd.Timestamp(existing_date).normalize() != pd.Timestamp(m_date).normalize()
                ):
                    exits_df.loc[idx, date_col] = m_date.strftime("%d-%b-%Y")
                    modified = True
                    synced  += 1
        if modified:
            exits_clean = exits_df.drop(columns=["_norm"], errors="ignore")
            with pd.ExcelWriter(
                self.excel_path, engine="openpyxl",
                mode="a", if_sheet_exists="replace"
            ) as writer:
                exits_clean.to_excel(writer, sheet_name=self.exits_sheet, index=False)
        return synced

    def _write_output(self, mismatches, unprocessed_exits, today_str):
        os.makedirs(self.out_folder, exist_ok=True)
        out_path = os.path.join(self.out_folder, "To_Be_Updated.xlsx")
        if os.path.exists(out_path):
            try:
                wb = load_workbook(out_path)
            except Exception:
                wb = Workbook()
                if "Sheet" in wb.sheetnames:
                    del wb["Sheet"]
        else:
            wb = Workbook()
            if "Sheet" in wb.sheetnames:
                del wb["Sheet"]

        if mismatches:
            self._write_sheet(wb, "Date Mismatches", self._MISMATCH_COLS,
                              mismatches, _TEAL, _AMBER_BG,
                              highlight_col="Date to Update To",
                              highlight_bg=_GREEN_BG)
        if unprocessed_exits:
            self._write_sheet(wb, "Unprocessed Exits", self._EXIT_COLS,
                              unprocessed_exits, _RED, _RED_BG)
        wb.save(out_path)
        if self.ui:
            self.ui.success(f"Output written → {out_path}")

    def _write_sheet(self, wb, sheet_name, cols, rows,
                     hdr_bg, row_bg, highlight_col=None, highlight_bg=None):
        if sheet_name in wb.sheetnames:
            ws       = wb[sheet_name]
            next_row = ws.max_row + 1
            if next_row == 1:
                self._write_header(ws, cols, hdr_bg, next_row)
                next_row += 1
        else:
            ws = wb.create_sheet(sheet_name)
            self._write_header(ws, cols, hdr_bg, 1)
            ws.row_dimensions[1].height = 28
            ws.freeze_panes = "A2"
            _autowidth(ws, cols)
            next_row = 2

        highlight_idx = (cols.index(highlight_col) + 1
                         if highlight_col and highlight_col in cols else None)

        for record in rows:
            for c_idx, col in enumerate(cols, 1):
                val  = record.get(col, "")
                cell = ws.cell(row=next_row, column=c_idx, value=val)
                bg   = (highlight_bg
                        if highlight_idx and c_idx == highlight_idx
                        else row_bg)
                _dat(cell, bg=bg)
            next_row += 1

    def _write_header(self, ws, cols, bg, row=1):
        for c_idx, col in enumerate(cols, 1):
            _hdr(ws.cell(row=row, column=c_idx, value=col), bg=bg)

    @staticmethod
    def _norm(name):
        return _norm_fn(name)