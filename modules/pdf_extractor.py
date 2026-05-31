"""
modules/pdf_extractor.py

Extracts employee data from medical PDF files received at BAC.
Handles three form types:
    Form 1 — EXIT MEDICAL EXAMINATION FORM
    Form 2 — MEDICAL CERTIFICATE OF FITNESS  (scanned, typed fields)
    Form 3 — CERTIFICATE OF FITNESS          (scanned, handwritten checkboxes)

Extraction priority:
    1. Name        → filename (most reliable, avoids handwriting problem)
    2. DateDone    → PDF text layer via form-aware regex (most critical field)
    3. MedicalType → PDF text layer (Periodic / Exit / Pre-Employment / Special)
    4. Other fields→ PDF text layer where available
    5. All fields  → headcount enrichment for anything still missing
    6. Failures    → flagged to "Manual Review" sheet in Excel workbook
"""

import os
import re
import pandas as pd
import pdfplumber
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config import CONFIG

# ── Date formats accepted across all three form types ─────────
# Covers: DD/MM/YYYY  DD-MM-YYYY  YYYY-MM-DD  DD Mon YYYY  D Mon YYYY
_DATE_FORMATS = [
    "%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d",
    "%d %B %Y", "%d %b %Y",
    "%d-%b-%Y", "%d-%b-%y",   # e.g. 30-Jan-2026
    "%d/%m/%y", "%d-%m-%y",
]

# ── Compiled regex patterns per form type ─────────────────────

# Generic date value — matches most written date formats
_RE_DATE_VALUE = (
    r"(\d{1,2}[\/\-]\d{1,2}[\/\-]\d{2,4}"          # DD/MM/YYYY or DD-MM-YY
    r"|\d{4}[\/\-]\d{2}[\/\-]\d{2}"                  # YYYY-MM-DD
    r"|\d{1,2}[\-\s](?:Jan|Feb|Mar|Apr|May|Jun|"
    r"Jul|Aug|Sep|Oct|Nov|Dec)[a-z]*[\-\.\s]\d{4})"  # DD-Mon-YYYY or DD Mon YYYY
)
# ── Form 1: EXIT MEDICAL EXAMINATION FORM ─────────────────────
FORM1_PATTERNS = {
    # "Date:" sits at the very top of Form 1
    "DateDone":    rf"(?:^|\n)\s*Date\s*[:\.]?\s*{_RE_DATE_VALUE}",

    # Personal details block
    "FirstName":   r"First\s*Name\s+([A-Za-z][A-Za-z '\-]{1,40}?)(?:\s{2,}|\n|Initials)",
    "Surname":     r"Surname\s+([A-Za-z][A-Za-z '\-]{1,40}?)(?:\s{2,}|\n|Date)",
    "Gender":      r"Gender\s+([MF](?:ale|emale)?)",
    "IDNumber":    r"ID\s*Number\s+([A-Z0-9]{6,20})",
    "JobTitle":    r"Job\s*Title\s+([A-Za-z][A-Za-z &\/\-]{2,50}?)(?:\s{2,}|\n|Work)",
    "WorkArea":    r"Work\s*Area\s+([A-Za-z][A-Za-z &\/\-]{2,50}?)(?:\s{2,}|\n|Exposure)",
    "Department":  r"(?:Department|Dept)[:\.]?\s+([A-Za-z][A-Za-z &\/\-]{2,50}?)(?:\s{2,}|\n)",
    "MedicalType": r"(Exit\s*Medical)",               # Form 1 is always Exit
}

# ── Form 2: MEDICAL CERTIFICATE OF FITNESS ────────────────────
FORM2_PATTERNS = {
    # "Date of issue" is the exam date per your confirmation
    "DateDone":    rf"Date\s*of\s*[Ii]ssue\s*[:\.]?\s*{_RE_DATE_VALUE}",

    # Name sits on its own line after the "Name" label
    "FullName":    r"Name\s*\n?\s*([A-Za-z][A-Za-z '\-]{2,60}?)(?:\n|Security)",
    "IDNumber":    r"Security\s*[Nn]umber\s*[:\.]?\s*([A-Z0-9]{6,20})",
    "JobTitle":    r"Job\s*[Ss]pecification\s*[:\.]?\s*([A-Za-z][A-Za-z &\/\-]{2,50}?)(?:\n|Clinic|Date)",
    "Department":  r"Department\s*[:\.]?\s*([A-Za-z][A-Za-z &\/\-]{2,50}?)(?:\n|Type)",
    "Clinic":      r"Clinic\s*[:\.]?\s*([A-Za-z][A-Za-z &\/\-., ]{2,60}?)(?:\n|Date)",

    # Type of medical — listed as checkboxes; pdfplumber picks up the label
    # text even when the box itself is handwritten, because the labels are printed
    "MedicalType": (
        r"(?:Medicals\s*[-–>]+\s*)"
        r"(Periodic|Pre[\s\-]*[Ee]mployment|Exit)"
        r"|(?:^|\n)\s*(Periodic|Pre[\s\-]*[Ee]mployment|Exit)\s*(?:\n|$)"
    ),
}

# ── Form 3: CERTIFICATE OF FITNESS ────────────────────────────
FORM3_PATTERNS = {
    # "Date: _" at the top
    "DateDone":    rf"Date\s*[:\._]{{0,3}}\s*{_RE_DATE_VALUE}",

    # "To: __" contains the employee name or department reference
    "ToField":     r"To\s*[:\._]{0,3}\s*\n?\s*([A-Za-z][A-Za-z '\-,]{2,80}?)(?:\n|of ID)",

    # ID / Passport number line
    "IDNumber":    r"of\s*ID\s*[\/\\]\s*Passport\s*No\s*[:\._]?\s*([A-Z0-9 \-]{6,25})",

    # Checkbox labels that appear near ticked boxes
    # pdfplumber extracts ALL printed text including unticked labels,
    # so we look for the label closest to a tick marker (✓ ✗ X / or similar)
    # Strategy: find whichever medical type label appears with a tick nearby
    "MedicalType": (
        r"(?:[\✓✗Xx\/\\]\s*)"                         # tick/cross marker
        r"(Pre[\s\-]*[Ee]mployment\s*Medicals?"
        r"|Periodical?\s*Medicals?"
        r"|Special\s*Medicals?"
        r"|Exit\s*Medicals?)"
        r"|"                                            # OR label appears alone on a line
        r"(?:^|\n)\s*"
        r"(Pre[\s\-]*[Ee]mployment\s*Medicals?"
        r"|Periodical?\s*Medicals?"
        r"|Special\s*Medicals?"
        r"|Exit\s*Medicals?)"
        r"\s*(?:\n|$)"
    ),

    # Certificate number
    "CertNo":      r"No\s*[:\._]?\s*([A-Z0-9\/\-]{3,20})",
}

# ── Canonical medical type mapping ────────────────────────────
# Normalises variations to a clean standard value
_MEDICAL_TYPE_MAP = {
    r"periodic":         "Periodic",
    r"pre.?employ":      "Pre-Employment",
    r"exit":             "Exit",
    r"special":          "Special",
    r"consultation":     "Consultation",
    r"biological":       "Biological Test",
    r"vaccination":      "Vaccination",
    r"drug.*alcohol":    "Drug/Alcohol Test",
    r"primary.*health":  "Primary Health Care",
}

# ── Filename noise words to strip when extracting name ────────
_NOISE_BASE = [
    # Form/document type words
    "HHS", "COF", "form", "forms",
    "certificate", "fitness", "medical", "medicals",
    "exam", "examination", "report", "results",
    # Medical type words
    "periodic", "periodical", "periodicals",
    "exit", "pre", "employment", "preemployment",
    "special", "consultation", "biological",
    "vaccination", "drug", "alcohol",
    # Generic filler
    "of", "the", "and", "for", "new",
]
# Merge with any extra words defined in config
_NOISE = list({*_NOISE_BASE, *CONFIG.get("FILENAME_NOISE", [])})


# ══════════════════════════════════════════════════════════════
class PDFExtractor:
    """
    Extracts structured data from medical PDF files.
    Automatically detects which of the three form types the PDF is.
    """

    MANUAL_REVIEW_SHEET = "Manual Review"

    def __init__(self, ui=None):
        self.ui = ui

    # ── Public entry point ────────────────────────────────────
    def extract(self, pdf_path):
        filename = os.path.basename(pdf_path)
        if self.ui:
            self.ui.info(f"Extracting: {filename}")

        extracted = {"SourceFile": filename}

        # ── Step 1: Name from filename ────────────────────────
        name, name_confidence = self._name_from_filename(filename)
        if name:
            extracted["Personnel Names"] = name
            extracted["FallbackUsed"]    = "Filename"
            if self.ui:
                self.ui.success(f"Name from filename: {name} "
                                f"[confidence: {name_confidence}]")

        # ── Step 2: Read PDF text layer ───────────────────────
        text = self._read_pdf_text(pdf_path)
        has_text = bool(text and text.strip())

        if has_text:
            # Detect form type and apply matching patterns
            form_type = self._detect_form_type(text)
            extracted["FormType"] = form_type
            if self.ui:
                self.ui.info(f"Form type detected: {form_type}")

            patterns = self._get_patterns(form_type)
            self._apply_patterns(extracted, text, patterns, form_type)

            # Fallback name from PDF body if filename failed
            if not extracted.get("Personnel Names"):
                body_name = self._name_from_body(text, form_type)
                if body_name:
                    extracted["Personnel Names"] = body_name
                    extracted["FallbackUsed"]    = "PDF Body"
                    if self.ui:
                        self.ui.info(f"Name from PDF body: {body_name}")
        else:
            extracted["FormType"]    = "Unknown (no text layer)"
            extracted["FallbackUsed"] = extracted.get("FallbackUsed", "Filename only")
            if self.ui:
                self.ui.warn("No text layer found — scanned PDF, "
                             "name from filename only")

        # ── Step 3: Enrich from headcount ─────────────────────
        extracted = self._enrich_from_headcount(extracted)

        # ── Step 4: Validate — flag if critical fields missing ─
        needs_review, reasons = self._needs_manual_review(extracted)
        if needs_review:
            extracted["NeedsReview"] = True
            extracted["ReviewReason"] = "; ".join(reasons)
            self._flag_for_review(extracted, pdf_path)
            if self.ui:
                self.ui.warn(f"Flagged for manual review: {'; '.join(reasons)}")
        else:
            extracted["NeedsReview"] = False

        return extracted

    # ── Form type detection ───────────────────────────────────
    def _detect_form_type(self, text):
        """
        Identifies which of the three form types this PDF is.
        Uses distinctive header phrases present in each form.
        """
        t = text.upper()
        if "EXIT MEDICAL EXAMINATION" in t:
            return "Form1_ExitMedical"
        if "MEDICAL CERTIFICATE OF FITNESS" in t:
            return "Form2_MedCertFitness"
        if "CERTIFICATE OF FITNESS" in t:
            return "Form3_CertFitness"
        # Fallback — try to infer from content keywords
        if "SECURITY NUMBER" in t or "JOB SPECIFICATION" in t:
            return "Form2_MedCertFitness"
        if "EXPOSURE GROUP" in t or "REGISTRY ID" in t:
            return "Form1_ExitMedical"
        if "PERIODICAL MEDICALS" in t or "PRE-EMPLOYMENT MEDICALS" in t:
            return "Form3_CertFitness"
        return "Unknown"

    def _get_patterns(self, form_type):
        return {
            "Form1_ExitMedical":    FORM1_PATTERNS,
            "Form2_MedCertFitness": FORM2_PATTERNS,
            "Form3_CertFitness":    FORM3_PATTERNS,
        }.get(form_type, {**FORM1_PATTERNS, **FORM2_PATTERNS, **FORM3_PATTERNS})

    # ── Pattern application ───────────────────────────────────
    def _apply_patterns(self, extracted, text, patterns, form_type):
        for field, pattern in patterns.items():
            if field in extracted:
                continue  # don't overwrite already-found values
            try:
                m = re.search(pattern, text,
                              re.IGNORECASE | re.MULTILINE)
                if m:
                    # Take first non-None group
                    value = next((g for g in m.groups() if g), None)
                    if value:
                        value = value.strip()
                        if field == "DateDone":
                            parsed = self._parse_date(value)
                            if parsed:
                                extracted["DateDone"] = parsed
                        elif field == "MedicalType":
                            extracted["MedicalType"] = self._normalise_medical_type(value)
                        elif field in ("FirstName", "Surname"):
                            # Form 1 splits name — merge when both found
                            extracted[field] = value
                            if "FirstName" in extracted and "Surname" in extracted:
                                extracted.setdefault(
                                    "Personnel Names",
                                    f"{extracted['FirstName']} {extracted['Surname']}".strip()
                                )
                        elif field == "FullName":
                            extracted.setdefault("Personnel Names", value)
                        elif field == "ToField":
                            # Form 3 "To:" may contain a name or dept — store raw
                            extracted["ToField"] = value
                            # Use as name only if it looks like a person's name
                            if self._looks_like_name(value):
                                extracted.setdefault("Personnel Names", value)
                        else:
                            extracted[field] = value
            except re.error:
                continue

        # Form 1 always Exit — set if not already captured
        if form_type == "Form1_ExitMedical":
            extracted.setdefault("MedicalType", "Exit")

    # ── Date parsing ──────────────────────────────────────────
    def _parse_date(self, raw):
        """Tries every known date format, returns datetime or None."""
        raw = raw.strip().rstrip(".")
        for fmt in _DATE_FORMATS:
            try:
                return datetime.strptime(raw, fmt)
            except ValueError:
                continue
        # Last resort — let pandas try
        result = pd.to_datetime(raw, errors="coerce", dayfirst=True)
        return result if not pd.isna(result) else None

    # ── Medical type normalisation ────────────────────────────
    def _normalise_medical_type(self, raw):
        raw = raw.strip()
        for pattern, canonical in _MEDICAL_TYPE_MAP.items():
            if re.search(pattern, raw, re.IGNORECASE):
                return canonical
        return raw.title()

    # ── Name from filename ────────────────────────────────────
    def _name_from_filename(self, filename):
        """
        Strips extension and noise words, then checks what remains.
        Returns (name_or_None, confidence_string).

        Confidence levels:
            high   — at least 2 capitalised alpha words, no digits
            medium — 2+ words but contains digits or unusual chars
            low    — only 1 word after cleaning
            none   — nothing useful left
        """
        stem = os.path.splitext(filename)[0]

        # Replace underscores/hyphens/dots with spaces
        stem = re.sub(r"[_\-\.]+", " ", stem)

        # Strip noise words — multi-pass until no more changes
        prev = None
        while prev != stem:
            prev = stem
            for word in _NOISE:
                stem = re.sub(rf"\b{re.escape(word)}\b", " ", stem,
                              flags=re.IGNORECASE)

        # Strip standalone digits
        stem = re.sub(r"\b\d+\b", " ", stem)
        stem = re.sub(r"\s+", " ", stem).strip()

        # Only keep words that look like name parts (letters, apostrophe, hyphen)
        # This drops leftover noise like "Certificate", "Periodic" etc.
        all_words   = stem.split()
        alpha_words = [w for w in all_words
                       if re.match(r"^[A-Za-z][A-Za-z'\-]*$", w)]

        if len(alpha_words) == 0:
            return None, "none"

        if len(alpha_words) == 1:
            # Single word — not reliable as a standalone name
            return None, "low"

        # Two or more clean alpha words — high confidence name
        name = " ".join(w.capitalize() for w in alpha_words)
        return name, "high"

    # ── Name from PDF body ────────────────────────────────────
    def _name_from_body(self, text, form_type):
        """
        Targeted name extraction from PDF body as last resort.
        Uses form-type-specific strategies.
        """
        if form_type == "Form1_ExitMedical":
            # Try First Name + Surname fields separately
            fn = re.search(r"First\s*Name\s+([A-Za-z][A-Za-z '\-]{1,30})",
                           text, re.IGNORECASE)
            sn = re.search(r"Surname\s+([A-Za-z][A-Za-z '\-]{1,30})",
                           text, re.IGNORECASE)
            first = fn.group(1).strip() if fn else ""
            last  = sn.group(1).strip() if sn else ""
            full  = f"{first} {last}".strip()
            return full if len(full) > 3 else None

        if form_type == "Form2_MedCertFitness":
            # "Name" label followed by value on same or next line
            m = re.search(
                r"(?:^|\n)\s*Name\s*\n?\s*([A-Za-z][A-Za-z '\-]{2,50})",
                text, re.IGNORECASE | re.MULTILINE
            )
            return m.group(1).strip() if m else None

        if form_type == "Form3_CertFitness":
            # "To:" field may contain a name
            m = re.search(
                r"To\s*[:\._]{0,3}\s*\n?\s*([A-Za-z][A-Za-z '\-,]{2,60})",
                text, re.IGNORECASE
            )
            if m:
                candidate = m.group(1).strip().split("\n")[0]
                if self._looks_like_name(candidate):
                    return candidate
        return None

    # ── Helpers ───────────────────────────────────────────────
    def _looks_like_name(self, text):
        """
        Heuristic: a name has 2–4 words, all alphabetic,
        no keywords that suggest it's a department or address.
        """
        NON_NAME_WORDS = {
            "the", "department", "manager", "director", "hr",
            "finance", "ict", "registry", "operations", "clinic",
            "gaborone", "maun", "francistown", "botswana",
        }
        words = text.strip().split()
        if not (2 <= len(words) <= 4):
            return False
        if any(not re.match(r"^[A-Za-z'\-]+$", w) for w in words):
            return False
        if any(w.lower() in NON_NAME_WORDS for w in words):
            return False
        return True

    def _read_pdf_text(self, pdf_path):
        try:
            with pdfplumber.open(pdf_path) as pdf:
                return "\n".join(
                    p.extract_text() or "" for p in pdf.pages
                )
        except Exception as e:
            if self.ui:
                self.ui.warn(f"pdfplumber error: {e}")
            return ""

    # ── Validation ────────────────────────────────────────────
    def _needs_manual_review(self, extracted):
        """
        Returns (True, [reasons]) if critical fields are missing
        or extraction confidence was too low to trust.
        """
        reasons = []

        if not extracted.get("Personnel Names"):
            reasons.append("Name could not be extracted")

        if not extracted.get("DateDone"):
            reasons.append("DateDone could not be extracted")

        # Name looks suspicious — too short or all caps (common OCR noise)
        name = extracted.get("Personnel Names", "")
        if name and (len(name) < 5 or name == name.upper()):
            reasons.append(f"Name suspect: '{name}'")

        return bool(reasons), reasons

    # ── Headcount enrichment ──────────────────────────────────
    def _enrich_from_headcount(self, extracted):
        name = extracted.get("Personnel Names")
        if not name:
            return extracted

        norm = self._normalise(name)
        hc_folder = CONFIG["HC_FOLDER"]

        try:
            files = sorted(
                [f for f in os.listdir(hc_folder)
                 if f.endswith((".xlsx", ".xls", ".csv"))],
                reverse=True   # most recent file first
            )
        except FileNotFoundError:
            return extracted

        for hc_file in files:
            try:
                path  = os.path.join(hc_folder, hc_file)
                hc_df = (pd.read_csv(path) if hc_file.endswith(".csv")
                         else pd.read_excel(path))

                if "Personnel Names" not in hc_df.columns:
                    continue

                hc_df["_norm"] = hc_df["Personnel Names"].apply(self._normalise)
                match = hc_df[hc_df["_norm"] == norm]

                if not match.empty:
                    row = match.iloc[0]
                    extracted.setdefault("EmployeeID", row.get("Pers.No."))
                    extracted.setdefault("PSGroup",    row.get("PS group"))
                    extracted.setdefault("Role",       row.get("Position"))
                    extracted.setdefault("Department", row.get("Personnel Area"))
                    extracted.setdefault("Gender",     row.get("Gender Key"))
                    if self.ui:
                        self.ui.success(f"Enriched from headcount: {name}")
                    return extracted
            except Exception:
                continue

        if self.ui:
            self.ui.warn(f"No headcount match for: {name}")
        return extracted

    # ── Manual review flagging ────────────────────────────────
    def _flag_for_review(self, extracted, pdf_path):
        """
        Writes a row to the 'Manual Review' sheet in the Excel workbook
        so someone can manually fill in the missing details.
        """
        excel_path = CONFIG["EXCEL_PATH"]
        sheet_name = self.MANUAL_REVIEW_SHEET

        record = {
            "Timestamp":       datetime.today().strftime("%d-%b-%Y %H:%M"),
            "SourceFile":      os.path.basename(pdf_path),
            "Personnel Names": extracted.get("Personnel Names", ""),
            "DateDone":        (extracted["DateDone"].strftime("%d-%b-%Y")
                                if extracted.get("DateDone") else ""),
            "MedicalType":     extracted.get("MedicalType", ""),
            "FormType":        extracted.get("FormType", ""),
            "ReviewReason":    extracted.get("ReviewReason", ""),
            "Action Taken":    "Pending",
        }

        try:
            wb = load_workbook(excel_path)

            if sheet_name not in wb.sheetnames:
                ws = wb.create_sheet(sheet_name)
                # Write header
                headers = list(record.keys())
                for c_idx, h in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=c_idx, value=h)
                    cell.font      = Font(bold=True, color="FFFFFF",
                                         name="Arial", size=10)
                    cell.fill      = PatternFill("solid", start_color="C00000")
                    cell.alignment = Alignment(horizontal="center",
                                               vertical="center",
                                               wrap_text=True)
                    cell.border    = self._thin_border()
                ws.row_dimensions[1].height = 28
                ws.freeze_panes = "A2"
                next_row = 2
            else:
                ws = wb[sheet_name]
                next_row = ws.max_row + 1

            # Write data row
            for c_idx, val in enumerate(record.values(), 1):
                cell = ws.cell(row=next_row, column=c_idx, value=val)
                cell.font      = Font(name="Arial", size=10)
                cell.alignment = Alignment(horizontal="left",
                                           vertical="center")
                cell.border    = self._thin_border()
                cell.fill      = PatternFill("solid", start_color="FCE4D6")

            # Auto-width
            for c_idx, (key, val) in enumerate(record.items(), 1):
                col_letter = get_column_letter(c_idx)
                current = ws.column_dimensions[col_letter].width or 0
                needed  = min(max(len(str(key)), len(str(val))) + 4, 50)
                if needed > current:
                    ws.column_dimensions[col_letter].width = needed

            wb.save(excel_path)

        except Exception as e:
            if self.ui:
                self.ui.error(f"Could not write to Manual Review sheet: {e}")

    @staticmethod
    def _thin_border():
        s = Side(style="thin")
        return Border(left=s, right=s, top=s, bottom=s)

    @staticmethod
    def _normalise(name):
        if pd.isna(name):
            return ""
        return re.sub(r"\s+", " ", str(name).strip().lower())
